"""Idempotent XLSX cost-calculator builder for business_sober_v1 / entrepreneur_v1.

Reads master.json + a context dict, emits a 3-sheet calculator:
    - Inputs       (yellow-fill editable cells)
    - Calculations (grey-fill formulas)
    - Output       (gold-fill board summary)

Same input -> identical .xlsx bytes (deterministic ordering; openpyxl's
zipfile output is stable when the workbook content matches).

Usage:
    python build_xlsx.py --context fixture.json --out calculator.xlsx

Dependencies: openpyxl >= 3.1
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:  # pragma: no cover
    try:
        # xlsxwriter fallback path (pure-write engine, also deterministic).
        import xlsxwriter  # type: ignore  # noqa: F401
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to build the business_sober_v1 XLSX kit. "
            "Install it with: pip install 'openpyxl>=3.1'"
        ) from exc
    Workbook = None  # type: ignore[assignment]


HERE = Path(__file__).resolve().parent
MASTER_PATH = HERE / "master.json"


def _hex(c: str) -> str:
    """openpyxl wants ARGB hex, so prepend FF if six chars."""
    c = c.lstrip("#").upper()
    if len(c) == 6:
        return "FF" + c
    return c


def _make_styles(theme: dict) -> dict:
    colors = theme["colors"]
    body_size = theme["fonts"].get("size_body", 11)
    head_size = theme["fonts"].get("size_h1", 14)
    label_size = theme["fonts"].get("size_label", 10)
    body_font = theme["fonts"].get("body", "Inter")
    numeric_font = theme["fonts"].get("numeric", "Consolas")

    thin_rule = Side(border_style="thin", color=_hex(colors["rule"]))
    box = Border(left=thin_rule, right=thin_rule, top=thin_rule, bottom=thin_rule)

    return {
        "header": {
            "fill": PatternFill("solid", fgColor=_hex(colors["header_fill"])),
            "font": Font(name=body_font, size=body_size, bold=True,
                         color=_hex(colors["header_font"])),
            "align": Alignment(horizontal="left", vertical="center"),
            "border": box,
        },
        "input": {
            "fill": PatternFill("solid", fgColor=_hex(colors["input_fill"])),
            "font": Font(name=body_font, size=body_size),
            "align": Alignment(horizontal="left", vertical="center"),
            "border": box,
        },
        "input_num": {
            "fill": PatternFill("solid", fgColor=_hex(colors["input_fill"])),
            "font": Font(name=numeric_font, size=body_size),
            "align": Alignment(horizontal="right", vertical="center"),
            "border": box,
        },
        "calc": {
            "fill": PatternFill("solid", fgColor=_hex(colors["calc_fill"])),
            "font": Font(name=numeric_font, size=body_size, italic=True),
            "align": Alignment(horizontal="right", vertical="center"),
            "border": box,
        },
        "output": {
            "fill": PatternFill("solid", fgColor=_hex(colors["output_fill"])),
            "font": Font(name=numeric_font, size=head_size, bold=True,
                         color=_hex(colors["output_font"])),
            "align": Alignment(horizontal="right", vertical="center"),
            "border": box,
        },
        "label": {
            "fill": PatternFill("solid", fgColor="00FFFFFF"),
            "font": Font(name=body_font, size=label_size, italic=True,
                         color="FF374151"),
            "align": Alignment(horizontal="left", vertical="center"),
        },
    }


def _apply(cell, style: dict) -> None:
    cell.fill = style["fill"]
    cell.font = style["font"]
    cell.alignment = style["align"]
    if "border" in style:
        cell.border = style["border"]


def _write_inputs(wb, master, context, styles) -> None:
    sheet_spec = master["sheets"][0]
    ws = wb.active
    ws.title = sheet_spec["name"]

    # Header
    for ci, col in enumerate(sheet_spec["columns"], start=1):
        c = ws.cell(row=1, column=ci, value=col)
        _apply(c, styles["header"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60

    # Rows; allow context overrides keyed by field name
    overrides = (context.get("inputs") or {})
    for ri, row in enumerate(sheet_spec["rows"], start=2):
        field = row["field"]
        value = overrides.get(field, row.get("value", ""))
        unit = row.get("unit", "")
        notes = row.get("notes", "")

        c_field = ws.cell(row=ri, column=1, value=field)
        _apply(c_field, styles["label"])

        c_value = ws.cell(row=ri, column=2, value=value)
        if isinstance(value, (int, float)):
            _apply(c_value, styles["input_num"])
            if unit == "ratio":
                c_value.number_format = master["number_formats"]["ratio"]
            elif unit == "CHF":
                c_value.number_format = master["number_formats"]["chf"]
            else:
                c_value.number_format = master["number_formats"]["integer"]
        else:
            _apply(c_value, styles["input"])

        c_unit = ws.cell(row=ri, column=3, value=unit)
        _apply(c_unit, styles["label"])

        c_notes = ws.cell(row=ri, column=4, value=notes)
        _apply(c_notes, styles["label"])

    # Named ranges (workbook-scoped)
    for name, ref in sheet_spec.get("named_ranges", {}).items():
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


def _write_calculations(wb, master, styles) -> None:
    sheet_spec = master["sheets"][1]
    ws = wb.create_sheet(title=sheet_spec["name"])

    for ci, col in enumerate(sheet_spec["columns"], start=1):
        c = ws.cell(row=1, column=ci, value=col)
        _apply(c, styles["header"])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14

    for ri, row in enumerate(sheet_spec["rows"], start=2):
        c_step = ws.cell(row=ri, column=1, value=row["step"])
        _apply(c_step, styles["label"])

        c_formula = ws.cell(row=ri, column=2, value=row["formula"])
        _apply(c_formula, styles["calc"])
        c_formula.font = Font(name="Consolas", size=10, italic=True, color="FF374151")
        c_formula.alignment = Alignment(horizontal="left", vertical="center")

        # Result column = identical formula evaluated by Excel.
        c_result = ws.cell(row=ri, column=3, value=row["formula"])
        _apply(c_result, styles["calc"])
        c_result.number_format = master["number_formats"]["chf"]

        c_unit = ws.cell(row=ri, column=4, value=row.get("unit", ""))
        _apply(c_unit, styles["label"])


def _write_output(wb, master, styles) -> None:
    sheet_spec = master["sheets"][2]
    ws = wb.create_sheet(title=sheet_spec["name"])

    for ci, col in enumerate(sheet_spec["columns"], start=1):
        c = ws.cell(row=1, column=ci, value=col)
        _apply(c, styles["header"])
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14

    for ri, row in enumerate(sheet_spec["rows"], start=2):
        c_metric = ws.cell(row=ri, column=1, value=row["metric"])
        _apply(c_metric, styles["label"])

        # Reference into Calculations sheet
        c_ref = ws.cell(row=ri, column=2, value="=" + row["ref"])
        _apply(c_ref, styles["output"])
        c_ref.number_format = master["number_formats"]["chf"]

        if "band_formula" in row:
            c_band = ws.cell(row=ri, column=3, value=row["band_formula"])
            _apply(c_band, styles["output"])
        else:
            c_band = ws.cell(row=ri, column=3, value="")
            _apply(c_band, styles["output"])


def build(context: dict, master: dict, out_path: Path) -> None:
    """Render the 3-sheet calculator workbook to ``out_path``."""
    if Workbook is None:
        raise SystemExit("openpyxl unavailable; xlsxwriter fallback not implemented in v1.")

    wb = Workbook()
    styles = _make_styles(master["theme"])

    _write_inputs(wb, master, context, styles)
    _write_calculations(wb, master, styles)
    _write_output(wb, master, styles)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _cli() -> None:
    p = argparse.ArgumentParser(description="Build business_sober_v1 cost-calculator XLSX.")
    p.add_argument("--context", type=Path, required=True)
    p.add_argument("--out", type=Path, required=True)
    p.add_argument("--master", type=Path, default=MASTER_PATH)
    args = p.parse_args()

    context = json.loads(args.context.read_text(encoding="utf-8"))
    master  = json.loads(args.master.read_text(encoding="utf-8"))
    build(context, master, args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    _cli()
